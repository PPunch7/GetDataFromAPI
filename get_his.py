import urllib.request, json
import sys
import datetime
import openpyxl

def checkNone(value):
    if value is None:
        return -999.99
    else:
        return value
    
def dt_format(value):
    dt = value.split('T')
    date = dt[0]
    time = dt[1].split('+')[0]
    return date + ' ' + time
    
target_dir = [DIR_NAME]
target_file = [FILE_NAME]
url = [LINK]
now = datetime.datetime.now()

# # file name
# now = datetime.datetime.now()
# filename = "his_" + start_date_obj.strftime("%Y%m%d") + "_" + now.strftime("%Y%m%d-%H%M%S%f") + ".txt"

# # CREATE file
# f = open(filename, "w")
# f.write("dam_code,site_timestamp,volume,inflow,outflow/n")
# f.close()

# OPEN result file
wb = openpyxl.load_workbook(filename=target_dir + target_file)

# GET JSON
response = urllib.request.urlopen(url)
print(url)

data = json.loads(response.read())

# print(data['data'])

old_sta = ''
# go through the DATA
for d in data['data']:
    #print (d)

    # get each DATA
    sta_id = d['stn_id']
    dt = dt_format(d['datetimes'])
    tmp = float(d['temp'])
    con = float(d['conducted'])
    tds = float(d['tds'])
    sal = float(d['salinity'])
    deo = float(d['deo'])
    dep = float(d['depth'])
    ph = float(d['ph'])
    nh4 = float(d['nh4'])
    tur = float(d['turbid'])

    print(sta_id, dt, tmp, con, tds, sal, deo, dep, ph, nh4, tur)

    # check sheet (sta_id) is exist
    if sta_id in wb.sheetnames:
        ws = wb[sta_id]
    else:
        wb.create_sheet(sta_id)
        ws = wb[sta_id]
        ws.cell(row=1, column=1).value = 'DATETIME'
        ws.cell(row=1, column=2).value = 'TEMP'
        ws.cell(row=1, column=3).value = 'CONDUCTED'
        ws.cell(row=1, column=4).value = 'TDS'
        ws.cell(row=1, column=5).value = 'SALINITY'
        ws.cell(row=1, column=6).value = 'DEO'
        ws.cell(row=1, column=7).value = 'DEPTH'
        ws.cell(row=1, column=8).value = 'pH'
        ws.cell(row=1, column=9).value = 'NH4'
        ws.cell(row=1, column=10).value = 'TURBID'
        wb.save(target_file)

    if sta_id != old_sta:
        # find EMPTY row
        target_row = 2
        while ws.cell(row=target_row, column=1).value != None:
            target_row += 1
    else:
        target_row += 1

    # define ALL value of row
    ws.cell(row=target_row, column=1).value = dt
    ws.cell(row=target_row, column=2).value = tmp
    ws.cell(row=target_row, column=3).value = con
    ws.cell(row=target_row, column=4).value = tds
    ws.cell(row=target_row, column=5).value = sal
    ws.cell(row=target_row, column=6).value = deo
    ws.cell(row=target_row, column=7).value = dep
    ws.cell(row=target_row, column=8).value = ph
    ws.cell(row=target_row, column=9).value = nh4
    ws.cell(row=target_row, column=10).value = tur

    

# # go through REGION
# for r in data['data']:
#     # print(r)
    
#     # go through DAM
#     for d in r['dam']:
#         # print(d)
#         id = d['id']
#         vol = checkNone(d['volume'])
#         inf = checkNone(d['inflow'])
#         outf = checkNone(d['outflow'])
#         # print (id, date_selected + ' ' + fix_t, vol, inf, outf)

#         # WRITE DATA
#         f = open(filename, "a")
#         f.write(id + "," + date_selected + ' ' + fix_t + "," + str(vol) + "," + str(inf) + "," + str(outf) + "/n")

# start_date_obj = start_date_obj + datetime.timedelta(days=1)
wb.save(target_file)  
wb.close()